import os
from openpyxl import load_workbook
from PIL import Image as PILImage
from openpyxl.drawing.image import Image

# Путь к папке
folder_path = "exel"
file_list = os.listdir(folder_path)

for file in file_list:
    if file.endswith(".xlsx"):

        # Путь к Excel файлу
        xlsx_file_path = f'{folder_path}/{file}'

        print(xlsx_file_path)

        # Путь к папке с изображениями
        img_folder_path = 'photo/'

        # Загрузка файла
        workbook = load_workbook(xlsx_file_path)
        sheet = workbook.active

        # Начальная строка
        start_row = 5
        # start_row = 9

        actual_row = start_row

        # Индекс столбца со значениями
        string_column_index = 1

        # Индекс столбца для вставки изображений
        image_column_index = 2

        for row in sheet.iter_rows(min_row=start_row, min_col=string_column_index, values_only=True):
            cell_value = row[3]  # Получаем значение ячейки


            # Генерируем путь к изображению
            image_path = os.path.join(img_folder_path, f'{cell_value}.JPG')


            # Проверяем, существует ли файл изображения в папке с исходниками
            if os.path.exists(image_path):
                print(image_path)
                img = PILImage.open(image_path)
                img_xlsx = Image(img)
                # Вставляем изображение
                sheet.add_image(img_xlsx, f"{chr(64 + image_column_index)}{actual_row}")

            else:
                pass

            actual_row += 1
            # break

        # Сохраняем изменения в Excel файле
        workbook.save(f'{folder_path}/add_photo_{file}')
